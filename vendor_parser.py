"""
Vendor xlsx parser — turns the RAD Global Vendor Database workbook into
structured dicts ready for db.merge_vendor_upload().

Three sheets are in scope (the other 3 in the source workbook are ignored):
  * Vendors           → list of vendor dicts (TEXT[] lists split from `;`)
  * RFP Categories    → list of category dicts (keywords split from `;`)
  * Products by Vendor → list of product dicts; reconciled to vendors by company name

Headers are detected by name (case-insensitive), so column reorderings or
extra columns are tolerated. Missing optional columns become empty defaults
without a warning; missing required columns surface as row-level warnings
and the row is skipped.

Public surface:
    parse_workbook(xlsx_bytes: bytes) -> ParseResult

`ParseResult` exposes `.vendors`, `.categories`, `.products`, `.warnings`.
"""
from __future__ import annotations

import io
import logging
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Any, Iterable, Optional

log = logging.getLogger(__name__)

# ── Public types ─────────────────────────────────────────────────────────────


@dataclass
class ParseResult:
    vendors: list[dict] = field(default_factory=list)
    categories: list[dict] = field(default_factory=list)
    products: list[dict] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


class VendorParseError(Exception):
    """Raised when the xlsx itself is unusable (not zip / no sheets)."""


# ── Sheet column maps (canonical → header substrings to match) ───────────────
# Match is case-insensitive on exact equality first, then substring.

VENDOR_COLS = {
    "company":          ("company",),
    "domain":           ("domain",),
    "primary_contacts": ("primary contact",),
    "emails":           ("email",),
    "phones":           ("phone",),
    "websites":         ("website",),
    "inquiry_count":    ("inquiry count",),
    "last_contact":     ("last contact",),
    "rfp_categories":   ("rfp categor",),
    "products_quoted":  ("products quoted",),
    "bid_count":        ("bid count",),
    "bid_history":      ("bid history",),
    "rfps_won":         ("rfps won",),
}

CATEGORY_COLS = {
    "category":                 ("category",),
    "keywords":                 ("keyword",),
    "vendors_tagged_count":     ("vendors tagged",),
    "needs_enrichment_count":   ("needs enrichment",),
}

PRODUCT_COLS = {
    "vendor_company":     ("vendor",),
    "domain_sender_key":  ("domain", "sender key"),
    "product":            ("product",),
    "rfp_code":           ("rfp code",),
    "source_tab":         ("source tab",),
}

VENDOR_REQUIRED = ("company",)
CATEGORY_REQUIRED = ("category",)
PRODUCT_REQUIRED = ("vendor_company", "product")


# ── Helpers ─────────────────────────────────────────────────────────────────

def _build_column_index(header_row: Iterable[Any], cols_map: dict) -> dict[str, int]:
    """Map canonical column name → 0-based index in the header row.

    Matching strategy:
      1. exact case-insensitive equality
      2. substring (any candidate substring present in the header cell)

    Columns not found are simply absent from the returned dict.
    """
    norm_headers = [(i, (str(h).strip().lower() if h is not None else ""))
                    for i, h in enumerate(header_row)]
    out: dict[str, int] = {}
    for canonical, candidates in cols_map.items():
        candidates_lc = tuple(c.lower() for c in candidates)
        # Exact match first
        for i, h in norm_headers:
            if h in candidates_lc:
                out[canonical] = i
                break
        else:
            # Substring fallback
            for i, h in norm_headers:
                if any(c in h for c in candidates_lc):
                    out[canonical] = i
                    break
    return out


def _cell(row: tuple, idx_map: dict[str, int], key: str) -> Any:
    """Get a cell value by canonical column name; returns None if column absent
    or row shorter than the column index."""
    idx = idx_map.get(key)
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _to_str(v: Any) -> str:
    """Cell → trimmed string. None / empty cells become ''."""
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    return str(v).strip()


def _split_semicolon_list(v: Any) -> list[str]:
    """Split a semicolon-delimited cell into a trimmed, de-empty'd list.

    Handles common author drift: trailing semicolons, double semicolons,
    spaces around items. Order is preserved.
    """
    s = _to_str(v)
    if not s:
        return []
    return [part for part in (p.strip() for p in s.split(";")) if part]


def _to_int(v: Any, *, default: int = 0) -> tuple[int, bool]:
    """Coerce a cell value to int. Returns (value, ok)."""
    if v is None or (isinstance(v, str) and not v.strip()):
        return default, True
    try:
        if isinstance(v, bool):  # bool is a subclass of int; reject
            return default, False
        return int(float(v)), True
    except (TypeError, ValueError):
        return default, False


_DATE_FORMATS = (
    "%Y-%m-%d",
    "%m/%d/%Y",
    "%d/%m/%Y",
    "%Y/%m/%d",
    "%m-%d-%Y",
)


def _to_date(v: Any) -> tuple[Optional[date], bool]:
    """Coerce a cell value to date. Returns (date or None, ok).

    Accepts datetime, date, common string formats. Empty cells → (None, True).
    """
    if v is None or (isinstance(v, str) and not v.strip()):
        return None, True
    if isinstance(v, datetime):
        return v.date(), True
    if isinstance(v, date):
        return v, True
    s = str(v).strip()
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt).date(), True
        except ValueError:
            continue
    return None, False


def _is_blank_row(row: tuple) -> bool:
    """A row is blank if every cell is None or an empty/whitespace string."""
    return all(c is None or (isinstance(c, str) and not c.strip()) for c in row)


# ── Per-sheet parsers ────────────────────────────────────────────────────────

def _parse_vendors_sheet(ws) -> tuple[list[dict], list[str]]:
    warnings: list[str] = []
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        warnings.append("Vendors: sheet is empty.")
        return [], warnings

    header = rows[0]
    idx_map = _build_column_index(header, VENDOR_COLS)
    missing_required = [k for k in VENDOR_REQUIRED if k not in idx_map]
    if missing_required:
        warnings.append(
            f"Vendors: required column(s) missing: {', '.join(missing_required)}. Sheet skipped."
        )
        return [], warnings

    out: list[dict] = []
    for row_no, row in enumerate(rows[1:], start=2):
        if _is_blank_row(row):
            continue

        company = _to_str(_cell(row, idx_map, "company"))
        if not company:
            warnings.append(f"Vendors row {row_no}: missing Company. Row skipped.")
            continue

        inq_count, ok_iq = _to_int(_cell(row, idx_map, "inquiry_count"))
        if not ok_iq:
            warnings.append(f"Vendors row {row_no} ({company}): Inquiry Count not numeric. Set to 0.")

        bid_count, ok_bid = _to_int(_cell(row, idx_map, "bid_count"))
        if not ok_bid:
            warnings.append(f"Vendors row {row_no} ({company}): Bid Count not numeric. Set to 0.")

        last_contact, ok_dt = _to_date(_cell(row, idx_map, "last_contact"))
        if not ok_dt:
            warnings.append(f"Vendors row {row_no} ({company}): Last Contact unparseable. Set to null.")

        out.append({
            "company":           company,
            "domain":            _to_str(_cell(row, idx_map, "domain")),
            "primary_contacts":  _split_semicolon_list(_cell(row, idx_map, "primary_contacts")),
            "emails":            _split_semicolon_list(_cell(row, idx_map, "emails")),
            "phones":            _split_semicolon_list(_cell(row, idx_map, "phones")),
            "websites":          _split_semicolon_list(_cell(row, idx_map, "websites")),
            "inquiry_count":     inq_count,
            "last_contact":      last_contact,
            "rfp_categories":    _split_semicolon_list(_cell(row, idx_map, "rfp_categories")),
            "products_quoted":   _split_semicolon_list(_cell(row, idx_map, "products_quoted")),
            "bid_count":         bid_count,
            "bid_history":       _to_str(_cell(row, idx_map, "bid_history")),
            "rfps_won":          _to_str(_cell(row, idx_map, "rfps_won")),
        })

    return out, warnings


def _parse_categories_sheet(ws) -> tuple[list[dict], list[str]]:
    warnings: list[str] = []
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        warnings.append("RFP Categories: sheet is empty.")
        return [], warnings

    header = rows[0]
    idx_map = _build_column_index(header, CATEGORY_COLS)
    missing_required = [k for k in CATEGORY_REQUIRED if k not in idx_map]
    if missing_required:
        warnings.append(
            f"RFP Categories: required column(s) missing: {', '.join(missing_required)}. Sheet skipped."
        )
        return [], warnings

    out: list[dict] = []
    for row_no, row in enumerate(rows[1:], start=2):
        if _is_blank_row(row):
            continue

        category = _to_str(_cell(row, idx_map, "category"))
        if not category:
            warnings.append(f"RFP Categories row {row_no}: missing Category. Row skipped.")
            continue

        vt, ok_vt = _to_int(_cell(row, idx_map, "vendors_tagged_count"))
        if not ok_vt:
            warnings.append(f"RFP Categories row {row_no} ({category}): Vendors tagged not numeric.")
        ne, ok_ne = _to_int(_cell(row, idx_map, "needs_enrichment_count"))
        if not ok_ne:
            warnings.append(f"RFP Categories row {row_no} ({category}): Needs Enrichment not numeric.")

        out.append({
            "category":               category,
            "keywords":               _split_semicolon_list(_cell(row, idx_map, "keywords")),
            "vendors_tagged_count":   vt,
            "needs_enrichment_count": ne,
        })
    return out, warnings


def _parse_products_sheet(ws) -> tuple[list[dict], list[str]]:
    warnings: list[str] = []
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        warnings.append("Products by Vendor: sheet is empty.")
        return [], warnings

    header = rows[0]
    idx_map = _build_column_index(header, PRODUCT_COLS)
    missing_required = [k for k in PRODUCT_REQUIRED if k not in idx_map]
    if missing_required:
        warnings.append(
            f"Products by Vendor: required column(s) missing: {', '.join(missing_required)}. Sheet skipped."
        )
        return [], warnings

    out: list[dict] = []
    for row_no, row in enumerate(rows[1:], start=2):
        if _is_blank_row(row):
            continue

        vendor_company = _to_str(_cell(row, idx_map, "vendor_company"))
        product = _to_str(_cell(row, idx_map, "product"))
        if not vendor_company or not product:
            warnings.append(f"Products by Vendor row {row_no}: missing Vendor or Product. Row skipped.")
            continue

        out.append({
            "vendor_company":    vendor_company,
            "domain_sender_key": _to_str(_cell(row, idx_map, "domain_sender_key")),
            "product":           product,
            "rfp_code":          _to_str(_cell(row, idx_map, "rfp_code")),
            "source_tab":        _to_str(_cell(row, idx_map, "source_tab")) or "Products by Vendor",
        })
    return out, warnings


def _reconcile_products(
    vendors: list[dict],
    products: list[dict],
) -> list[str]:
    """Surface a warning for each product whose vendor doesn't exist in the
    Vendors sheet. The DB layer will still insert the row (vendor_id NULL)
    so the data isn't lost — this is informational."""
    known = {(v.get("company") or "").strip().lower() for v in vendors}
    orphans: dict[str, int] = {}
    for p in products:
        company = (p.get("vendor_company") or "").strip().lower()
        if company and company not in known:
            orphans[company] = orphans.get(company, 0) + 1
    warnings: list[str] = []
    for company, n in sorted(orphans.items()):
        warnings.append(
            f"Orphan product: vendor '{company}' appears in Products by Vendor "
            f"({n} row{'s' if n != 1 else ''}) but not in the Vendors sheet."
        )
    return warnings


# ── Public entry point ──────────────────────────────────────────────────────

def parse_workbook(xlsx_bytes: bytes) -> ParseResult:
    """Parse the in-scope sheets of a RAD Vendor Database xlsx workbook."""
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover
        raise VendorParseError("openpyxl is required for vendor uploads") from exc

    if not xlsx_bytes:
        raise VendorParseError("Empty file.")

    try:
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    except Exception as exc:
        raise VendorParseError(f"Could not open workbook: {exc}") from exc

    result = ParseResult()

    # Sheet lookup is case-insensitive; tolerate minor naming differences.
    sheet_lookup = {name.strip().lower(): name for name in wb.sheetnames}

    def _find(*candidates: str) -> Optional[str]:
        for c in candidates:
            real = sheet_lookup.get(c.lower())
            if real:
                return real
        return None

    vendors_sheet = _find("Vendors")
    categories_sheet = _find("RFP Categories", "RFP Category", "Categories")
    products_sheet = _find("Products by Vendor", "Products By Vendor", "Products")

    if vendors_sheet is None:
        result.warnings.append("Workbook is missing the 'Vendors' sheet.")
    else:
        vendors, w = _parse_vendors_sheet(wb[vendors_sheet])
        result.vendors = vendors
        result.warnings.extend(w)

    if categories_sheet is None:
        result.warnings.append("Workbook is missing the 'RFP Categories' sheet.")
    else:
        cats, w = _parse_categories_sheet(wb[categories_sheet])
        result.categories = cats
        result.warnings.extend(w)

    if products_sheet is None:
        result.warnings.append("Workbook is missing the 'Products by Vendor' sheet.")
    else:
        prods, w = _parse_products_sheet(wb[products_sheet])
        result.products = prods
        result.warnings.extend(w)

    # Cross-sheet reconcile
    if result.vendors and result.products:
        result.warnings.extend(_reconcile_products(result.vendors, result.products))

    log.info(
        "vendor.parse.ok vendors=%d categories=%d products=%d warnings=%d",
        len(result.vendors), len(result.categories),
        len(result.products), len(result.warnings),
    )
    return result
